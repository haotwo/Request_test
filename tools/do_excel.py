#作者:HP
#日期:2020-03-18 21:53
#文件:do_excel
from openpyxl import load_workbook
from tools.read_config import ReadConfig
from tools import project_path
#相对路径  绝对路径
class DoExcel:
    @staticmethod
    def get_data(file_name):
        wb = load_workbook(file_name)
        mode=eval(ReadConfig.get_config(project_path.case_config_path,'MODE','mode'))

        # name="self:261993005056/123"
        test_data = []
        for key in mode:  #遍历这个存在配置文件里面的字典
            sheet = wb[key]   #表单名
            if mode[key]=='all':
                for i in range(2,sheet.max_row+1):
                    row_data={}   #字典
                    row_data['case_id']=sheet.cell(i,1).value
                    row_data['url']=sheet.cell(i,2).value
                    row_data['data']=sheet.cell(i,3).value
                    # if sheet.cell(i,3).value.find('${name}')!=-1:   #有找到这个${name}
                    #     row_data['data']=sheet.cell(i,3).value.replace('${name}',str(name))
                    # else:  #如果没有找到的话
                    #     row_data['data']=sheet.cell(i,3).vlaue
                    row_data['title']=sheet.cell(i,4).value
                    row_data['http_method']=sheet.cell(i,5).value
                    row_data['expected']=sheet.cell(i,6).value
                    row_data['sheet_name']=key
                    test_data.append(row_data)
            else:
                for case_id in mode[key]:
                    row_data = {}  # 字典
                    row_data['case_id'] = sheet.cell(case_id+1,1).value
                    row_data['url'] = sheet.cell(case_id+1,2).value
                    row_data['data'] = sheet.cell(case_id+1,3).value
                    row_data['title'] = sheet.cell(case_id+1,4).value
                    row_data['http_method'] = sheet.cell(case_id+1,5).value
                    row_data['expected'] = sheet.cell(case_id+1,6).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
        return test_data
    @staticmethod
    def write_back(file_name,sheet_name,i,result,TestrResult):   #写回数据
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(i,7).value=result
        sheet.cell(i,8).value=TestrResult
        wb.save(file_name)    #保存结果

    def updare_name(self,file_name,sheet_name,name):
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(2,1).value=name
        wb.save(file_name)
if __name__=='__main__':
    test_data=DoExcel().get_data(project_path.test_case_path)
    print(test_data)


